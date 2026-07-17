from __future__ import annotations

import io
from pathlib import Path
from typing import Any

import pandas as pd
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, Field

from analytics import build_dashboard, get_metadata, get_quality

BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"

app = FastAPI(
    title="Sistema de análisis, proyección y alerta de matrícula escolar en Junín",
    version="1.0.0",
)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class Filters(BaseModel):
    province: str | None = None
    district: str | None = None
    ugel: str | None = None
    levels: list[str] = Field(default_factory=list)
    managements: list[str] = Field(default_factory=list)
    areas: list[str] = Field(default_factory=list)


class DashboardRequest(BaseModel):
    start_year: int = 2021
    end_year: int = 2024
    comparison_group: str = "provincia"
    alert_group: str = "provincia"
    filters: Filters = Field(default_factory=Filters)


@app.get("/api/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/api/meta")
def metadata() -> dict[str, Any]:
    return get_metadata()


@app.get("/api/quality")
def quality() -> dict[str, Any]:
    return get_quality()


@app.post("/api/dashboard")
def dashboard(request: DashboardRequest) -> dict[str, Any]:
    try:
        return build_dashboard(request.model_dump())
    except Exception as exc:  # pragma: no cover - surfaced to frontend
        raise HTTPException(status_code=500, detail=f"No se pudo calcular el análisis: {exc}") from exc


def _autosize(ws) -> None:
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 48)
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="17324D")
        cell.alignment = Alignment(horizontal="center")


@app.post("/api/export/excel")
def export_excel(request: DashboardRequest) -> StreamingResponse:
    result = build_dashboard(request.model_dump())
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary = result["summary"]
        summary_rows = [
            ("Ámbito", summary["context"]),
            ("Año inicial", summary["start_year"]),
            ("Año final", summary["end_year"]),
            ("Matrícula inicial", summary["start_total"]),
            ("Matrícula final", summary["end_total"]),
            ("Cambio absoluto", summary["absolute_change"]),
            ("Cambio porcentual", summary["percentage_change"]),
            ("Proyección 2025", summary["forecast_2025"]),
            ("Proyección 2026", summary["forecast_2026"]),
            ("Territorios con atención", summary["attention_count"]),
            ("Territorios con demanda creciente", summary["growth_count"]),
        ]
        pd.DataFrame(summary_rows, columns=["Indicador", "Resultado"]).to_excel(writer, sheet_name="Resumen", index=False)
        pd.DataFrame(result["evolution"]["overall"]).to_excel(writer, sheet_name="Evolucion", index=False)
        pd.DataFrame(result["comparison"]["rows"]).to_excel(writer, sheet_name="Comparacion", index=False)
        pd.DataFrame(result["alerts"]["rows"]).drop(columns=["series"], errors="ignore").to_excel(writer, sheet_name="Alertas", index=False)
        projection_rows = result["projection"]["observed"] + [
            {"year": x["year"], "value": x["value"], "lower": x["lower"], "upper": x["upper"]}
            for x in result["projection"]["forecast"]
        ]
        pd.DataFrame(projection_rows).to_excel(writer, sheet_name="Proyeccion", index=False)
        quality_data = get_quality()
        pd.DataFrame(quality_data["metrics"]).to_excel(writer, sheet_name="Calidad", index=False)
        for ws in writer.book.worksheets:
            _autosize(ws)

    output.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="informe_matricula_junin.xlsx"'}
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


if STATIC_DIR.exists():
    app.mount("/assets", StaticFiles(directory=STATIC_DIR / "assets"), name="assets")

    @app.get("/{full_path:path}")
    def spa(full_path: str):
        requested = STATIC_DIR / full_path
        if full_path and requested.exists() and requested.is_file():
            return FileResponse(requested)
        return FileResponse(STATIC_DIR / "index.html")
